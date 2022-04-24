import pandas as pd
import numpy as np
import csv
import openpyxl
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

df = pd.read_csv(r"C:\Users\Jorge Cantu\Downloads\SampleUpdated.csv",encoding='latin1')

wb = openpyxl.Workbook()
ws = wb.active

with open(r"C:\Users\Jorge Cantu\Downloads\SampleUpdated.csv") as f:
    reader = csv.reader(f, delimiter=':')
    for row in reader:
        ws.append(row)

wb.save('Sample.xlsx')

wb = load_workbook('Sample.xlsx')

Cost = float(input("What is your max Monthly Rent:"))

NumberofBedRooms = float(input("What is desired number of Bedrooms/Bathrooms(1-4): "))

Distance = float(input("How far away would you like to live(In miles): "))

profile_id_4_df = df[df['Monthly Rent'] <= Cost]

profile_id_4_df = profile_id_4_df[profile_id_4_df['Number of Bedrooms'] <= NumberofBedRooms]

profile_id_4_df = profile_id_4_df[profile_id_4_df['Distance to Texas A&M University (miles)'] <= Distance]

profile_id_4_df.to_excel(r'C:\Users\Jorge Cantu\Downloads\SampleOutputList.xlsx')

read_file = pd.read_excel (r"C:\Users\Jorge Cantu\Downloads\SampleOutputList.xlsx")
  
# Write the dataframe object
# into csv file

read_file.to_csv (r"C:\Users\Jorge Cantu\Downloads\Test.csv",index = None,header=True)
