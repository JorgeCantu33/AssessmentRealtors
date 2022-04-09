from typing import Final
import streamlit as st
import pandas as pd
import numpy as np
import csv
import openpyxl
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

st.set_page_config(
    layout="wide"
)

st.write("""
# Assessment Realtors
##### Welcome, we are AssessmentRealtors.com 
To begin please answer the questions on the sidebar located on the left side of the screen and set your aparment specifications.
""")

st.markdown("---")

st.sidebar.header('Apartment Specifications')

Cost = float(st.sidebar.number_input (
    "Enter the max Monthly Rent:", min_value=0, max_value=2000, step=50
))

NumberofBedRooms = float(st.sidebar.selectbox(
    "Enter your desired number of Bedrooms and Bathrooms:", ("1", "2", "3", "4")
))

Distance = float(st.sidebar.selectbox(
    "How far away would you like to live(In miles):", ("2.5", "5", "7.5", "10", "12.5", "15")
))

df = pd.read_csv(r"C:\Users\Jorge Cantu\Downloads\SampleUpdated.csv",encoding='latin1')

wb = openpyxl.Workbook()
ws = wb.active

with open(r"C:\Users\Jorge Cantu\Downloads\SampleUpdated.csv") as f:
    reader = csv.reader(f, delimiter=':')
    for row in reader:
        ws.append(row)

wb.save('Sample.xlsx')

wb = load_workbook('Sample.xlsx')

#Cost = float(input("What is your max Monthly Rent:"))

#NumberofBedRooms = float(input("What is desired number of Bedrooms/Bathrooms(1-4): "))

#Distance = float(input("How far away would you like to live(In miles): "))

profile_id_4_df = df[df['Monthly Rent'] <= Cost]

profile_id_4_df = profile_id_4_df[profile_id_4_df['Number of Bedrooms'] <= NumberofBedRooms]

profile_id_4_df = profile_id_4_df[profile_id_4_df['Distance to Texas A&M University (miles)'] <= Distance]

profile_id_4_df.to_excel(r'C:\Users\Jorge Cantu\Downloads\SampleOutputList.xlsx')

read_file = pd.read_excel (r"C:\Users\Jorge Cantu\Downloads\SampleOutputList.xlsx")
  
# Write the dataframe object
# into csv file

read_file.to_csv (r"C:\Users\Jorge Cantu\Downloads\Test.csv",index = None,header=True)

Final = pd.read_csv(r"C:\Users\Jorge Cantu\Downloads\Test.csv",encoding='latin1')

st.dataframe(Final)
